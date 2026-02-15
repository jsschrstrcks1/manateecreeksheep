#!/usr/bin/env python3
"""
Generate individual breeding page sheets for each sheep in an Excel workbook.

Each sheet contains:
- Header info (name, tag, sex, status, pen, DOB)
- Breed composition table (all breeds)
- Pedigree tree (up to great-great-grandparents)
- Inbreeding coefficient
- Weight data (actual, measured, projected)
- Offspring table

Weight projection formula: (girth^2 * length) / 300
"""

import json
import os
from datetime import datetime
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(SCRIPT_DIR, "..", "data", "flock_database.json")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "..", "data", "breeding_pages.xlsx")

# ── All breed names in standard order (matches owner's breeding pages) ──
ALL_BREEDS = [
    "American Blackbelly",
    "Awassi",
    "Babydoll",
    "Barbados Blackbelly",
    "Black Headed Dorper",
    "Cotswold",
    "Cracker",
    "East Friesian",
    "Gulf Coast Native",
    "Hampshire",
    "Jacob",
    "Karakul",
    "Katahdin",
    "Southdown",
    "St Augustine",
    "St Croix",
    "Suffolk",
    "Texel",
    "Tunis",
    "White Dorper",
    "Wiltshire Horn",
]

# ── Styles ─────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=16)
SECTION_FONT = Font(name="Calibri", bold=True, size=12)
LABEL_FONT = Font(name="Calibri", bold=True, size=10)
DATA_FONT = Font(name="Calibri", size=10)
SMALL_FONT = Font(name="Calibri", size=9, italic=True)

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT_WHITE = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
BREED_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
PEDIGREE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
WEIGHT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
OFFSPRING_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
INBRED_FILL = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def load_database():
    with open(DB_PATH) as f:
        return json.load(f)


def get_sheep_by_id(db, sheep_id):
    """Look up a sheep record by ID."""
    for s in db["sheep"]:
        if s["id"] == sheep_id:
            return s
    return None


def get_ancestors(db, sheep_id, depth=4, visited=None):
    """
    Build ancestor tree up to `depth` generations.
    Returns dict: {"sheep": record, "sire": {...}, "dam": {...}}
    Also tracks all ancestor IDs for inbreeding detection.
    """
    if visited is None:
        visited = set()
    if depth <= 0 or not sheep_id or sheep_id in visited:
        return None

    s = get_sheep_by_id(db, sheep_id)
    if not s:
        return None

    visited.add(sheep_id)
    return {
        "sheep": s,
        "sire": get_ancestors(db, s.get("sire_id"), depth - 1, visited.copy()),
        "dam": get_ancestors(db, s.get("dam_id"), depth - 1, visited.copy()),
    }


def collect_ancestor_ids(tree, generation=0):
    """
    Collect all ancestor IDs with their generation depth.
    Returns list of (id, generation) tuples.
    """
    if not tree or not tree.get("sheep"):
        return []
    result = [(tree["sheep"]["id"], generation)]
    result += collect_ancestor_ids(tree.get("sire"), generation + 1)
    result += collect_ancestor_ids(tree.get("dam"), generation + 1)
    return result


def calculate_inbreeding(db, sheep_record):
    """
    Calculate inbreeding coefficient by checking if any ancestor appears
    on BOTH the sire's side and dam's side of the pedigree.
    Uses Wright's method approximation for simple cases.
    Returns (coefficient_percent, list_of_common_ancestors).
    """
    sire_id = sheep_record.get("sire_id")
    dam_id = sheep_record.get("dam_id")

    if not sire_id or not dam_id:
        return 0.0, []

    # Build separate trees for sire and dam sides
    sire_tree = get_ancestors(db, sire_id, depth=5)
    dam_tree = get_ancestors(db, dam_id, depth=5)

    if not sire_tree or not dam_tree:
        return 0.0, []

    sire_ancestors = collect_ancestor_ids(sire_tree)
    dam_ancestors = collect_ancestor_ids(dam_tree)

    sire_ids = {aid for aid, _ in sire_ancestors}
    dam_ids = {aid for aid, _ in dam_ancestors}

    common = sire_ids & dam_ids

    if not common:
        return 0.0, []

    # Approximate inbreeding: for each common ancestor, F += (1/2)^(n1+n2+1)
    # where n1 = generations from sire to common ancestor
    # and n2 = generations from dam to common ancestor
    sire_gen = {aid: gen for aid, gen in sire_ancestors}
    dam_gen = {aid: gen for aid, gen in dam_ancestors}

    f = 0.0
    common_names = []
    for ancestor_id in common:
        n1 = sire_gen.get(ancestor_id, 0)
        n2 = dam_gen.get(ancestor_id, 0)
        contribution = (0.5) ** (n1 + n2 + 1)
        f += contribution
        ancestor = get_sheep_by_id(db, ancestor_id)
        name = ancestor["name"] if ancestor else ancestor_id
        common_names.append(name)

    return round(f * 100, 2), common_names


def calculate_projected_weight(girth, length):
    """Sheep weight estimation: (girth^2 * length) / 300."""
    if girth and length:
        return round((girth ** 2 * length) / 300, 1)
    return None


def safe_sheet_name(name, idx):
    """
    Excel sheet names: max 31 chars, no []:*?/\\ characters.
    Append index to ensure uniqueness.
    """
    clean = name.replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "")
    clean = clean.replace("[", "(").replace("]", ")").replace(":", "-")
    # Truncate to leave room for index suffix
    max_len = 28
    if len(clean) > max_len:
        clean = clean[:max_len]
    return f"{clean} ({idx})"


def write_sheep_sheet(wb, db, sheep_record, idx):
    """Write one sheep's breeding page as a worksheet."""
    sid = sheep_record["id"]
    name = sheep_record["name"]
    ws = wb.create_sheet(title=safe_sheet_name(name, idx))

    # Column widths
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 20

    row = 1

    # ── HEADER ─────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1, value=name)
    cell.font = HEADER_FONT_WHITE
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    row += 1

    # Basic info grid
    info_fields = [
        ("Tag", sheep_record.get("tag", "[none]")),
        ("Sex", sheep_record.get("sex", "unknown").title()),
        ("Status", sheep_record.get("status", "unknown").title()),
        ("Pen", sheep_record.get("pen", "[unknown]")),
        ("DOB", sheep_record.get("dob", "[unknown]")),
        ("DOB Approximate", "Yes" if sheep_record.get("dob_approximate", True) else "No"),
        ("Aliases", ", ".join(sheep_record.get("aliases", [])) or "[none]"),
        ("Confidence", sheep_record.get("confidence", "medium").title()),
    ]

    for label, value in info_fields:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=str(value)).font = DATA_FONT
        row += 1

    row += 1

    # ── BREED COMPOSITION ──────────────────────────────────────
    ws.merge_cells(f"A{row}:C{row}")
    cell = ws.cell(row=row, column=1, value="BREED COMPOSITION")
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    ws.cell(row=row, column=2).fill = SECTION_FILL
    ws.cell(row=row, column=3).fill = SECTION_FILL
    row += 1

    # Column headers
    for col, header in enumerate(["Breed", "%", "Coat Type"], 1):
        c = ws.cell(row=row, column=col, value=header)
        c.font = LABEL_FONT
        c.fill = BREED_FILL
        c.border = THIN_BORDER
    row += 1

    breed_pcts = sheep_record.get("breed_composition", {}).get("percentages", {})
    coat_type = sheep_record.get("breed_composition", {}).get("coat_type", "unknown")
    hair_pct = sheep_record.get("breed_composition", {}).get("hair_percentage", "?")

    total_pct = 0
    for breed_name in ALL_BREEDS:
        pct = breed_pcts.get(breed_name, 0)
        ws.cell(row=row, column=1, value=breed_name).font = DATA_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        pct_cell = ws.cell(row=row, column=2, value=pct if pct > 0 else "")
        pct_cell.font = DATA_FONT
        pct_cell.border = THIN_BORDER
        pct_cell.alignment = CENTER
        if pct > 0:
            pct_cell.number_format = "0.###"
            pct_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        ws.cell(row=row, column=3).border = THIN_BORDER
        total_pct += pct
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = LABEL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    total_cell = ws.cell(row=row, column=2, value=round(total_pct, 2))
    total_cell.font = LABEL_FONT
    total_cell.border = THIN_BORDER
    total_cell.alignment = CENTER
    if abs(total_pct - 100) > 0.5:
        total_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.cell(row=row, column=3).border = THIN_BORDER
    row += 1

    # Coat type summary
    ws.cell(row=row, column=1, value="Coat Type").font = LABEL_FONT
    ws.cell(row=row, column=2, value=coat_type.title()).font = DATA_FONT
    row += 1
    ws.cell(row=row, column=1, value="Hair %").font = LABEL_FONT
    ws.cell(row=row, column=2, value=f"{hair_pct}%").font = DATA_FONT
    row += 2

    # ── PEDIGREE ───────────────────────────────────────────────
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1, value="PEDIGREE")
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    for c in range(2, 8):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    row += 1

    # Column headers for pedigree
    ped_headers = ["Relationship", "Name", "ID", "Breed", "Sex", "Status"]
    for col, header in enumerate(ped_headers, 1):
        c = ws.cell(row=row, column=col, value=header)
        c.font = LABEL_FONT
        c.fill = PEDIGREE_FILL
        c.border = THIN_BORDER
    row += 1

    def write_pedigree_row(ws, row, label, sheep_id, indent=0):
        """Write one pedigree row. Returns next row."""
        prefix = "  " * indent
        record = get_sheep_by_id(db, sheep_id) if sheep_id else None

        ws.cell(row=row, column=1, value=f"{prefix}{label}").font = LABEL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER

        if record:
            ws.cell(row=row, column=2, value=record["name"]).font = DATA_FONT
            ws.cell(row=row, column=3, value=record["id"]).font = SMALL_FONT
            breed_pcts = record.get("breed_composition", {}).get("percentages", {})
            breed_str = ", ".join(f"{p}% {b}" for b, p in sorted(breed_pcts.items(), key=lambda x: -x[1]) if p > 0)
            ws.cell(row=row, column=4, value=breed_str or "[unknown]").font = SMALL_FONT
            ws.cell(row=row, column=4).alignment = LEFT
            ws.cell(row=row, column=5, value=record.get("sex", "?").title()).font = DATA_FONT
            ws.cell(row=row, column=6, value=record.get("status", "?").title()).font = DATA_FONT
        else:
            ws.cell(row=row, column=2, value=sheep_id or "[unknown]").font = DATA_FONT
            for c in range(3, 7):
                ws.cell(row=row, column=c, value="—").font = DATA_FONT

        for c in range(2, 7):
            ws.cell(row=row, column=c).border = THIN_BORDER

        return row + 1

    # Write pedigree tree
    sire_id = sheep_record.get("sire_id")
    dam_id = sheep_record.get("dam_id")

    row = write_pedigree_row(ws, row, "Sire", sire_id, 0)
    if sire_id:
        sire = get_sheep_by_id(db, sire_id)
        if sire:
            row = write_pedigree_row(ws, row, "Sire's Sire (Grand)", sire.get("sire_id"), 1)
            if sire.get("sire_id"):
                gs = get_sheep_by_id(db, sire["sire_id"])
                if gs:
                    row = write_pedigree_row(ws, row, "Great-Grand Sire", gs.get("sire_id"), 2)
                    row = write_pedigree_row(ws, row, "Great-Grand Dam", gs.get("dam_id"), 2)
            row = write_pedigree_row(ws, row, "Sire's Dam (Grand)", sire.get("dam_id"), 1)
            if sire.get("dam_id"):
                gd = get_sheep_by_id(db, sire["dam_id"])
                if gd:
                    row = write_pedigree_row(ws, row, "Great-Grand Sire", gd.get("sire_id"), 2)
                    row = write_pedigree_row(ws, row, "Great-Grand Dam", gd.get("dam_id"), 2)

    row = write_pedigree_row(ws, row, "Dam", dam_id, 0)
    if dam_id:
        dam = get_sheep_by_id(db, dam_id)
        if dam:
            row = write_pedigree_row(ws, row, "Dam's Sire (Grand)", dam.get("sire_id"), 1)
            if dam.get("sire_id"):
                gs = get_sheep_by_id(db, dam["sire_id"])
                if gs:
                    row = write_pedigree_row(ws, row, "Great-Grand Sire", gs.get("sire_id"), 2)
                    row = write_pedigree_row(ws, row, "Great-Grand Dam", gs.get("dam_id"), 2)
            row = write_pedigree_row(ws, row, "Dam's Dam (Grand)", dam.get("dam_id"), 1)
            if dam.get("dam_id"):
                gd = get_sheep_by_id(db, dam["dam_id"])
                if gd:
                    row = write_pedigree_row(ws, row, "Great-Grand Sire", gd.get("sire_id"), 2)
                    row = write_pedigree_row(ws, row, "Great-Grand Dam", gd.get("dam_id"), 2)

    row += 1

    # ── INBREEDING COEFFICIENT ─────────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws.cell(row=row, column=1, value="INBREEDING COEFFICIENT")
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    for c in range(2, 5):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    row += 1

    inbreeding_pct, common_ancestors = calculate_inbreeding(db, sheep_record)

    ws.cell(row=row, column=1, value="Coefficient").font = LABEL_FONT
    coeff_cell = ws.cell(row=row, column=2, value=f"{inbreeding_pct}%")
    coeff_cell.font = Font(name="Calibri", bold=True, size=12,
                           color="CC0000" if inbreeding_pct > 0 else "006600")
    if inbreeding_pct > 0:
        coeff_cell.fill = INBRED_FILL
    row += 1

    if common_ancestors:
        ws.cell(row=row, column=1, value="Common Ancestors").font = LABEL_FONT
        ws.cell(row=row, column=2, value=", ".join(common_ancestors)).font = DATA_FONT
        row += 1

    if inbreeding_pct == 0:
        ws.cell(row=row, column=1, value="No common ancestors detected in pedigree").font = SMALL_FONT
    elif inbreeding_pct > 12:
        ws.cell(row=row, column=1, value="⚠ HIGH INBREEDING — avoid breeding to related animals").font = Font(
            name="Calibri", bold=True, size=10, color="CC0000")
    row += 2

    # ── WEIGHT DATA ────────────────────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws.cell(row=row, column=1, value="WEIGHT DATA")
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    for c in range(2, 5):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    row += 1

    actual_weight = sheep_record.get("weight_lbs")
    measurements = sheep_record.get("measurements", {})
    girth = measurements.get("girth")
    length = measurements.get("length")
    calc_weight = measurements.get("calculated_weight")
    meas_date = measurements.get("date")

    projected = calculate_projected_weight(girth, length)

    weight_rows = [
        ("Actual Weight (lbs)", f"{actual_weight}" if actual_weight else "[not recorded]"),
        ("Girth (in)", f"{girth}" if girth else "[not measured]"),
        ("Length (in)", f"{length}" if length else "[not measured]"),
        ("Projected Weight (lbs)", f"{projected}" if projected else "[insufficient measurements]"),
        ("Formula", "girth² × length ÷ 300" if girth and length else ""),
        ("Measurement Date", str(meas_date) if meas_date else "[unknown]"),
    ]

    if calc_weight and projected:
        diff = abs(calc_weight - projected)
        if diff > 1:
            weight_rows.append(("DB Calculated Weight", f"{calc_weight} lbs"))

    for label, value in weight_rows:
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = LABEL_FONT
        c1.fill = WEIGHT_FILL
        c1.border = THIN_BORDER
        c2 = ws.cell(row=row, column=2, value=value)
        c2.font = DATA_FONT
        c2.border = THIN_BORDER
        row += 1

    # Weight comparison
    if actual_weight and projected:
        diff = actual_weight - projected
        ws.cell(row=row, column=1, value="Actual vs Projected").font = LABEL_FONT
        ws.cell(row=row, column=1).fill = WEIGHT_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        diff_str = f"{'+' if diff > 0 else ''}{round(diff, 1)} lbs ({'+' if diff > 0 else ''}{round(diff/projected*100, 1)}%)"
        ws.cell(row=row, column=2, value=diff_str).font = DATA_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    row += 1

    # ── OFFSPRING ──────────────────────────────────────────────
    offspring_ids = sheep_record.get("breeding", {}).get("offspring_ids", [])

    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1, value=f"OFFSPRING ({len(offspring_ids)})")
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    for c in range(2, 8):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    row += 1

    if offspring_ids:
        off_headers = ["Name", "ID", "Sex", "DOB", "Other Parent", "Status", "Breed"]
        for col, header in enumerate(off_headers, 1):
            c = ws.cell(row=row, column=col, value=header)
            c.font = LABEL_FONT
            c.fill = OFFSPRING_FILL
            c.border = THIN_BORDER
        row += 1

        for oid in offspring_ids:
            offspring = get_sheep_by_id(db, oid)
            if offspring:
                # Determine other parent
                if offspring.get("sire_id") == sid:
                    other_parent_id = offspring.get("dam_id")
                elif offspring.get("dam_id") == sid:
                    other_parent_id = offspring.get("sire_id")
                else:
                    other_parent_id = None

                other_parent = get_sheep_by_id(db, other_parent_id) if other_parent_id else None
                other_name = other_parent["name"] if other_parent else (other_parent_id or "[unknown]")

                breed_pcts = offspring.get("breed_composition", {}).get("percentages", {})
                breed_str = ", ".join(f"{p}%{b[:3]}" for b, p in sorted(breed_pcts.items(), key=lambda x: -x[1]) if p > 0)

                vals = [
                    offspring["name"],
                    offspring["id"],
                    offspring.get("sex", "?").title(),
                    offspring.get("dob", "[unknown]"),
                    other_name,
                    offspring.get("status", "?").title(),
                    breed_str or "[unknown]",
                ]
            else:
                vals = [oid, oid, "?", "?", "?", "?", "?"]

            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = DATA_FONT
                c.border = THIN_BORDER
                if col >= 4:
                    c.alignment = LEFT
            row += 1
    else:
        ws.cell(row=row, column=1, value="No offspring recorded").font = SMALL_FONT
        row += 1

    row += 1

    # ── MEDICAL / HEALTH HISTORY ──────────────────────────────
    health = sheep_record.get("health", {})
    famacha = health.get("famacha_scores", [])
    treatments = health.get("treatments", [])
    vaccinations = health.get("vaccinations", [])
    weak_resistance = health.get("weak_resistance", False)
    health_notes = health.get("notes", [])

    has_health_data = famacha or treatments or vaccinations or weak_resistance or health_notes

    MEDICAL_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1, value="MEDICAL / HEALTH HISTORY")
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    for c in range(2, 8):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    row += 1

    # Weak resistance flag
    ws.cell(row=row, column=1, value="Weak Resistance").font = LABEL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    wr_cell = ws.cell(row=row, column=2, value="YES — on weak resistance list" if weak_resistance else "No")
    wr_cell.font = Font(name="Calibri", bold=weak_resistance, size=10,
                        color="CC0000" if weak_resistance else "006600")
    wr_cell.border = THIN_BORDER
    if weak_resistance:
        wr_cell.fill = INBRED_FILL
    row += 1

    # FAMACHA scores
    if famacha:
        ws.cell(row=row, column=1, value="FAMACHA History").font = LABEL_FONT
        ws.cell(row=row, column=1).fill = MEDICAL_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        row += 1
        fam_headers = ["Date", "Score", "Notes"]
        for col, h in enumerate(fam_headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = LABEL_FONT
            c.border = THIN_BORDER
            c.fill = MEDICAL_FILL
        row += 1
        for entry in famacha:
            ws.cell(row=row, column=1, value=str(entry.get("date", ""))).font = DATA_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            score_cell = ws.cell(row=row, column=2, value=entry.get("score", ""))
            score_cell.font = DATA_FONT
            score_cell.border = THIN_BORDER
            score_cell.alignment = CENTER
            # Color code FAMACHA: 1=green, 2=light green, 3=yellow, 4=orange, 5=red
            score = entry.get("score", 0)
            try:
                score = int(score)
            except (ValueError, TypeError):
                score = 0
            if score <= 2:
                score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif score == 3:
                score_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            elif score >= 4:
                score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            ws.cell(row=row, column=3, value=str(entry.get("notes", ""))).font = DATA_FONT
            ws.cell(row=row, column=3).border = THIN_BORDER
            row += 1
        row += 1

    # Treatments
    if treatments:
        ws.cell(row=row, column=1, value="Treatment History").font = LABEL_FONT
        ws.cell(row=row, column=1).fill = MEDICAL_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        row += 1
        tx_headers = ["Date", "Treatment"]
        for col, h in enumerate(tx_headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = LABEL_FONT
            c.border = THIN_BORDER
            c.fill = MEDICAL_FILL
        row += 1
        for entry in treatments:
            ws.cell(row=row, column=1, value=str(entry.get("date", ""))).font = DATA_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=2, value=str(entry.get("treatment", ""))).font = DATA_FONT
            ws.cell(row=row, column=2).border = THIN_BORDER
            row += 1
        row += 1

    # Vaccinations
    if vaccinations:
        ws.cell(row=row, column=1, value="Vaccinations").font = LABEL_FONT
        ws.cell(row=row, column=1).fill = MEDICAL_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        row += 1
        for entry in vaccinations:
            ws.cell(row=row, column=1, value=str(entry.get("date", ""))).font = DATA_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=2, value=str(entry.get("vaccine", ""))).font = DATA_FONT
            ws.cell(row=row, column=2).border = THIN_BORDER
            row += 1
        row += 1

    if not has_health_data:
        ws.cell(row=row, column=1, value="No medical records on file").font = SMALL_FONT
        row += 1

    row += 1

    # ── NOTES ─────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1, value="NOTES")
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    for c in range(2, 8):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    row += 1

    notes = sheep_record.get("notes", "")
    if notes:
        ws.merge_cells(f"A{row}:G{row}")
        c = ws.cell(row=row, column=1, value=notes)
        c.font = DATA_FONT
        c.alignment = LEFT
        ws.row_dimensions[row].height = max(30, len(notes) // 3)
    else:
        ws.cell(row=row, column=1, value="[no notes]").font = SMALL_FONT

    return ws


def main():
    print("Loading flock database...")
    db = load_database()
    sheep_list = db["sheep"]

    # Sort: alive first, then by name
    status_order = {"alive": 0, "sold": 1, "unknown": 2, "deceased": 3, "culled": 4}
    sheep_list.sort(key=lambda s: (status_order.get(s["status"], 5), s["name"]))

    print(f"Generating breeding pages for {len(sheep_list)} sheep...")

    wb = Workbook()

    # ── INDEX SHEET ────────────────────────────────────────────
    ws_index = wb.active
    ws_index.title = "Index"
    ws_index.column_dimensions["A"].width = 6
    ws_index.column_dimensions["B"].width = 24
    ws_index.column_dimensions["C"].width = 10
    ws_index.column_dimensions["D"].width = 10
    ws_index.column_dimensions["E"].width = 10
    ws_index.column_dimensions["F"].width = 12
    ws_index.column_dimensions["G"].width = 40

    ws_index.merge_cells("A1:G1")
    c = ws_index.cell(row=1, column=1, value="Manatee Creek Sheep — Breeding Pages")
    c.font = HEADER_FONT_WHITE
    c.fill = HEADER_FILL
    c.alignment = CENTER

    ws_index.cell(row=2, column=1, value=f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = SMALL_FONT

    headers = ["#", "Name", "Tag", "Sex", "Status", "Pen", "Primary Breed"]
    for col, h in enumerate(headers, 1):
        c = ws_index.cell(row=4, column=col, value=h)
        c.font = LABEL_FONT
        c.fill = SECTION_FILL
        c.border = THIN_BORDER

    for i, s in enumerate(sheep_list):
        r = 5 + i
        primary = s.get("breed_composition", {}).get("primary", "[unknown]")
        vals = [i + 1, s["name"], s.get("tag", ""), s.get("sex", "?"), s.get("status", "?"), s.get("pen", ""), primary]
        for col, val in enumerate(vals, 1):
            c = ws_index.cell(row=r, column=col, value=val)
            c.font = DATA_FONT
            c.border = THIN_BORDER

    # ── INDIVIDUAL SHEETS ──────────────────────────────────────
    for i, s in enumerate(sheep_list):
        write_sheep_sheet(wb, db, s, i + 1)
        if (i + 1) % 20 == 0:
            print(f"  ... {i + 1}/{len(sheep_list)} sheets created")

    print(f"  ... {len(sheep_list)}/{len(sheep_list)} sheets created")

    # Save
    wb.save(OUTPUT_PATH)
    print(f"\nBreeding pages saved to: {OUTPUT_PATH}")
    print(f"  Sheets: {len(wb.sheetnames)} (1 index + {len(sheep_list)} individual)")


if __name__ == "__main__":
    main()
